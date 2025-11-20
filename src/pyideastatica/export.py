import os
import time
import json
from openpyxl import Workbook
import ideastatica_connection_api
import ideastatica_connection_api.connection_api_service_attacher as connection_api_service_attacher
from ideastatica_connection_api.models.con_load_effect import ConLoadEffect
from ideastatica_connection_api.models.con_calculation_parameter import ConCalculationParameter
from ideastatica_connection_api.models.con_connection import ConConnection

BASE_URL = "http://localhost:5000"  # Default, do not touch


# def validate_fatigue_analysis_type(
#     connection: ConConnection, stress_strain: bool = True, fatigue: bool = False
# ) -> None:
#     if stress_strain:
#         if connection.analysis_type != "stress_Strain":
#             raise ValueError("Connection type is not Stress-Strain analysis! Please change your file.")
#     if fatigue:
#         if connection.analysis_type != "fatigues":
#             raise ValueError("Connection type is not Fatigue analysis! Please change your file.")


def get_connection_by_name(api_client, project_id, connection_name):
    all_connections = api_client.connection.get_connections(project_id)
    for con in all_connections:
        if con.name == connection_name:
            return con
    # If reach this point, nothing matched
    raise ValueError(f"Connection with name '{connection_name}' not found.")


def export_weld_stress(file_path, connection_name, write_json, output_path):
    with connection_api_service_attacher.ConnectionApiServiceAttacher(BASE_URL).create_api_client() as api_client:
        uploadRes = api_client.project.open_project_from_filepath(file_path)  # Is this needed?
        project_id = api_client.project.active_project_id
        # project_data = api_client.project.get_project_data(project_id)
        connection = get_connection_by_name(
            api_client=api_client, project_id=project_id, connection_name=connection_name
        )
        file_name = os.path.basename(file_path)
        if connection.analysis_type == "stress_Strain":
            _export_weld_stress_strain(api_client, connection_name, write_json, file_name, output_path)
        elif connection.analysis_type == "fatigues":
            _export_weld_stress_fatigue(api_client, connection_name, write_json, file_name, output_path)
        else:
            raise ValueError(f"Connection  '{connection_name}' analysis type not supported.")


def _export_weld_stress_strain(api_client, connection_name, write_json, file_name, output_path):
    """Exports weld stresses to excel, assuming stress-strain analysis."""
    start_time = time.time()
    project_id = api_client.project.active_project_id

    # Get connection in the project
    connection1 = get_connection_by_name(api_client=api_client, project_id=project_id, connection_name=connection_name)

    # Get all loads
    loads = api_client.load_effect.get_load_effects(project_id, connection1.id)

    no_loads = len(loads)
    loads_copy = loads

    output = []

    for n in range(no_loads):
        loop_loads = api_client.load_effect.get_load_effects(project_id, connection1.id)
        # print("load count at the start of loop: " + str(len(loop_loads)))
        load_ef = loop_loads[n]
        current_name = load_ef.name
        # print("Processing case:.... " + str(current_name))
        print(f"Processing case:....{current_name} ({n+1}/{no_loads})")
        # Remove other loads than one currently considered
        for load in loop_loads:
            if load.name != current_name:
                api_client.load_effect.delete_load_effect(project_id, connection1.id, load.id)
                # print("Deleted case: " + str(load.name))

        # currrent_loads = api_client.load_effect.get_load_effects(project_id, connection1.id)
        # print(len(currrent_loads))

        calcParams = ideastatica_connection_api.ConCalculationParameter()
        calcParams.connection_ids = [connection1.id]

        # Model must be calculated first
        con1_cbfem_results = api_client.calculation.calculate(
            api_client.project.active_project_id, calcParams.connection_ids
        )

        # print(con1_cbfem_results)

        # detailed_results = api_client.calculation.get_results(
        #     api_client.project.active_project_id, calcParams.connection_ids
        # )
        # print(detailed_results)

        results_text = api_client.calculation.get_raw_json_results(
            api_client.project.active_project_id, calcParams
        )  # Does it return anything else than index 0?
        firstConnectionResult = results_text[0]

        raw_results = json.loads(firstConnectionResult)
        weld_results = raw_results["welds"]

        # Stressess are in N/m2 (Pa) so require *1e-6
        for weld_id, weld in weld_results.items():
            lcase = weld.get("loadCase")
            jname = weld.get("joinedItemName")
            name = weld.get("name")
            thickness = weld.get("thickness")
            design_thickness = weld.get("designedThickness")
            weld_type = weld.get("weldType2")
            weld_length = weld.get("length")
            sigma_per = weld.get("sigmaPerpendicular") * 1e-6
            tau_y = weld.get("tauy") * 1e-6
            tau_x = weld.get("taux") * 1e-6
            output.append(
                [lcase, jname, name, thickness, design_thickness, weld_type, weld_length, sigma_per, tau_y, tau_x]
            )

        # print(output)

        # Add all loads back again
        for copied_load in loads_copy:
            if copied_load.name != current_name:
                api_client.load_effect.add_load_effect(project_id, connection1.id, con_load_effect=copied_load)

    print("Saving data...")
    # Preamble for output data
    preamble = [
        "Load Case",
        "joinedItemName",
        "Name",
        "Thickness",
        "Design thickness",
        "Weld type",
        "Weld length",
        "\u03c3_\u27c2",  # sigma_perpendicular
        "\u03c4_\u27c2",  # tau_perpendicular
        "\u03c4_\u2225",  # tau_parallel
    ]
    output.insert(0, preamble)
    save_name = connection1.name

    timestr = time.strftime("%Y-%m-%d %H:%M:%S")
    info_output = [
        ["File name:", file_name],
        ["Connection name:", connection1.name],
        ["Date and time:", timestr],
        ["Analysis type:", connection1.analysis_type.value],
    ]

    # Write to excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Welds"
    ws3 = wb.create_sheet("Info")
    # Append rows to the worksheet
    for row in output:
        ws.append(row)
    for row in info_output:
        ws3.append(row)
    # Save the file

    excel_name = f"\\{save_name}_weld_stress.xlsx"
    json_name = f"\\{save_name}_data.json"
    out_path = output_path + excel_name
    json_out_path = output_path + json_name
    wb.save(out_path)

    if write_json:
        # Writing to the .json
        with open(json_out_path, "w") as f:
            json.dump(raw_results, f)

    elapsed = time.time() - start_time
    formatted = time.strftime("%H:%M:%S", time.gmtime(elapsed))
    print(f"Export complete in {formatted} sec.")


def _export_weld_stress_fatigue(api_client, connection_name, write_json, file_name, output_path):
    start_time = time.time()

    project_id = api_client.project.active_project_id
    # Get connection in the project
    connection1 = get_connection_by_name(api_client=api_client, project_id=project_id, connection_name=connection_name)

    # Get all loads
    loads = api_client.load_effect.get_load_effects(project_id, connection1.id)
    no_loads = len(loads)

    loads_copy = loads

    fatigue_output = []
    weld_output = []

    for n in range(1, no_loads):
        loop_loads = api_client.load_effect.get_load_effects(project_id, connection1.id)
        # print("Load cases count at the start of loop: " + str(len(loop_loads)))
        load_ef = loop_loads[n]
        current_name = load_ef.name
        print(f"Processing case:....{current_name} ({n}/{no_loads-1})")
        # Remove other loads than one currently considered
        for load in loop_loads:
            if load.name != current_name and load.name != "Ref_0MPa":
                api_client.load_effect.delete_load_effect(project_id, connection1.id, load.id)
                # print("Deleted case: " + str(load.name))

        currrent_loads = api_client.load_effect.get_load_effects(project_id, connection1.id)
        # print("Count of cases left in iteration: " + str(len(currrent_loads)))

        calcParams = ideastatica_connection_api.ConCalculationParameter()
        calcParams.connection_ids = [connection1.id]
        calcParams.analysis_type = "fatigues"
        # connection1.analysis_type = calcParams.analysis_type
        # updated_connection1 = api_client.connection.update_connection(
        #     api_client.project.active_project_id, connection1.id, connection1
        # )

        # Model must be calculated first
        con1_fatigue_results = api_client.calculation.calculate(
            api_client.project.active_project_id, calcParams.connection_ids
        )

        # print(con1_fatigue_results)

        # detailed_results = api_client.calculation.get_results(
        #     api_client.project.active_project_id, calcParams.connection_ids
        # )
        # print(detailed_results)

        results_text = api_client.calculation.get_raw_json_results(api_client.project.active_project_id, calcParams)
        firstConnectionResult = results_text[0]
        raw_results = json.loads(firstConnectionResult)

        fatigue_results = raw_results["fatigueChecks"]
        # Writing to the .json
        # print(weld_results)

        # Stressess are in N/m2 (Pa) so require *1e-6
        for weld_id, weld in fatigue_results.items():
            lcase = weld.get("loadCase")
            jname = weld.get("joinedItemName")
            thickness = weld.get("designedThickness")
            leg_size = weld.get("legSize")
            name = weld.get("name")
            weld_type = weld.get("weldType2")
            weld_length = weld.get("length")
            sigma_max = weld.get("normalStress") * 1e-6
            tau = weld.get("shearStress") * 1e-6
            sigma = weld.get("normalStress2") * 1e-6
            tau_max = weld.get("shearStress2") * 1e-6
            fatigue_output.append(
                [lcase, jname, name, thickness, leg_size, weld_type, weld_length, sigma_max, tau, tau_max, sigma]
            )

        weld_results = raw_results["fatigueWelds"]

        for weld_id, weld in weld_results.items():
            lcase = weld.get("loadCase")
            jname = weld.get("joinedItemName")
            name = weld.get("name")
            thickness = weld.get("designedThickness")
            leg_size = weld.get("legSize")
            weld_type = weld.get("weldType2")
            weld_length = weld.get("length")
            max_eq_stress = weld.get("maxEquivalentStress") * 1e-6
            tau_y = weld.get("tauy") * 1e-6
            tau_x = weld.get("taux") * 1e-6
            tau_wf_max = weld.get("tauxwf") * 1e-6
            sigma_wf = weld.get("sigmawf") * 1e-6
            weld_output.append(
                [
                    lcase,
                    jname,
                    name,
                    thickness,
                    leg_size,
                    weld_type,
                    weld_length,
                    max_eq_stress,
                    tau_y,
                    tau_x,
                    tau_wf_max,
                    sigma_wf,
                ]
            )

        # Add all loads back again
        for copied_load in loads_copy:
            if copied_load.name != current_name and copied_load.name != "Ref_0MPa":
                api_client.load_effect.add_load_effect(project_id, connection1.id, con_load_effect=copied_load)

    print("Saving data...")

    # Preamble for output data
    preamble_fatigue = [
        "Load Case",
        "Joined Item Name",
        "Name",
        "Designed Thickness",
        "Leg size",
        "Weld type",
        "Weld length",
        "Normal stress",
        "Shear stress",
        "Shear stress 2",
        "Normal stress 2",
    ]
    preamble_weld = [
        "Load Case",
        "Joined Item Name",
        "Name",
        "Designed Thickness",
        "Leg size",
        "Weld type",
        "Weld length",
        "Max \u03c3_eq stress",
        "\u03c4_y",
        "\u03c4_x",
        "\u03c4_wf_max",
        "\u03c3_wf",
    ]
    fatigue_output.insert(0, preamble_fatigue)
    weld_output.insert(0, preamble_weld)

    save_name = connection1.name

    timestr = time.strftime("%Y-%m-%d %H:%M:%S")
    info_output = [
        ["File name:", file_name],
        ["Connection name:", connection1.name],
        ["Date and time:", timestr],
        ["Analysis type:", connection1.analysis_type.value],
    ]

    # # # Write to .txt
    # # with open("weld_data.txt", "w") as f:
    # #     for line in output:
    # #         f.write(f"{line}\n")

    # Write to excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Welds"

    ws2 = wb.create_sheet("Weld sections")
    ws3 = wb.create_sheet("Info")

    # Append rows to the worksheet
    for row in weld_output:
        ws.append(row)

    for row in fatigue_output:
        ws2.append(row)

    for row in info_output:
        ws3.append(row)

    # Save the file
    excel_name = f"\\{save_name}_fatigue_weld_stress.xlsx"
    json_name = f"\\{save_name}_data.json"
    out_path = output_path + excel_name
    json_out_path = output_path + json_name
    wb.save(out_path)

    if write_json:
        with open(json_out_path, "w") as f:
            json.dump(raw_results, f)

    elapsed = time.time() - start_time
    formatted = time.strftime("%H:%M:%S", time.gmtime(elapsed))
    print(f"Export complete in {formatted} sec.")
